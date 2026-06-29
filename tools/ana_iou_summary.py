#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Summarize per-image analysis folders into one IoU comparison workbook.

Expected layout:
    ana/
      cxup_1b_BW/
        per_image_miou.csv
      unet_BW/
        per_image_miou.csv

Example:
    python tools/ana_iou_summary.py ana \
        --class-file classes.txt \
        --output ana/iou_summary.xlsx
"""

from __future__ import annotations

import argparse
import csv
import math
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable, Optional


CLASS_IOU_RE = re.compile(r"^class_(\d+)_iou$")
CLASS_INTERSECT_RE = re.compile(r"^class_(\d+)_intersect$")
CLASS_PRED_RE = re.compile(r"^class_(\d+)_pred_area$")
CLASS_LABEL_RE = re.compile(r"^class_(\d+)_label_area$")


@dataclass
class ParseIssue:
    source: str
    level: str
    message: str


@dataclass
class ClassSummary:
    class_id: int
    class_name: str
    iou: Optional[float] = None
    mean_image_iou: Optional[float] = None
    intersect: int = 0
    pred_area: int = 0
    label_area: int = 0
    union: int = 0


@dataclass
class ModelSummary:
    model_name: str
    csv_file: str
    csv_path: str
    num_images: int
    miou: Optional[float] = None
    mean_image_miou: Optional[float] = None
    class_metrics: list[ClassSummary] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


def read_text_auto(path: Path) -> str:
    for encoding in ("utf-8", "utf-8-sig", "gb18030", "latin-1"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    return path.read_text(encoding="utf-8", errors="replace")


def load_class_names(path: Optional[Path]) -> tuple[list[str], list[ParseIssue]]:
    if path is None:
        return [], []
    if not path.is_file():
        return [], [ParseIssue(path.name, "ERROR", f"Class file does not exist: {path}")]

    names = []
    for raw_line in read_text_auto(path).splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue

        # Accept "0: Background", "0,Background", "0 Background", etc.
        match = re.match(r"^\s*\d+\s*(?:[:;,|\t]|\s)\s*(.+?)\s*$", line)
        if match:
            line = match.group(1).strip()
        names.append(line)
    return names, []


def safe_float(value) -> Optional[float]:
    if value is None:
        return None
    text = str(value).strip()
    if text == "":
        return None
    try:
        result = float(text)
    except ValueError:
        return None
    if math.isnan(result) or math.isinf(result):
        return None
    return result


def safe_int(value) -> int:
    number = safe_float(value)
    return int(number) if number is not None else 0


def collect_class_ids(fieldnames: Iterable[str]) -> list[int]:
    ids = set()
    for name in fieldnames:
        for pattern in (
            CLASS_IOU_RE,
            CLASS_INTERSECT_RE,
            CLASS_PRED_RE,
            CLASS_LABEL_RE,
        ):
            match = pattern.match(name)
            if match:
                ids.add(int(match.group(1)))
    return sorted(ids)


def class_name_for(class_id: int, class_names: list[str]) -> str:
    if 0 <= class_id < len(class_names):
        return class_names[class_id]
    return f"class_{class_id}"


def parse_per_image_csv(path: Path, class_names: list[str]) -> tuple[Optional[ModelSummary], list[ParseIssue]]:
    issues: list[ParseIssue] = []
    model_name = path.parent.name

    with path.open("r", newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames or []
        rows = list(reader)

    if not rows:
        issues.append(ParseIssue(str(path), "ERROR", "per_image_miou.csv is empty"))
        return None, issues

    class_ids = collect_class_ids(fieldnames)
    if not class_ids:
        issues.append(ParseIssue(str(path), "ERROR", "No class metric columns found"))
        return None, issues

    intersections = {class_id: 0 for class_id in class_ids}
    pred_areas = {class_id: 0 for class_id in class_ids}
    label_areas = {class_id: 0 for class_id in class_ids}
    image_ious = {class_id: [] for class_id in class_ids}
    image_mious = []

    has_area_columns = all(
        f"class_{class_id}_intersect" in fieldnames
        and f"class_{class_id}_pred_area" in fieldnames
        and f"class_{class_id}_label_area" in fieldnames
        for class_id in class_ids
    )
    if not has_area_columns:
        issues.append(
            ParseIssue(
                str(path),
                "WARNING",
                "Area columns are incomplete; class IoU falls back to mean image IoU.",
            )
        )

    for row in rows:
        miou = safe_float(row.get("miou"))
        if miou is not None:
            image_mious.append(miou)

        for class_id in class_ids:
            image_iou = safe_float(row.get(f"class_{class_id}_iou"))
            if image_iou is not None:
                image_ious[class_id].append(image_iou)

            intersections[class_id] += safe_int(row.get(f"class_{class_id}_intersect"))
            pred_areas[class_id] += safe_int(row.get(f"class_{class_id}_pred_area"))
            label_areas[class_id] += safe_int(row.get(f"class_{class_id}_label_area"))

    class_metrics = []
    global_ious = []
    for class_id in class_ids:
        intersect = intersections[class_id]
        pred_area = pred_areas[class_id]
        label_area = label_areas[class_id]
        union = pred_area + label_area - intersect

        if has_area_columns:
            class_iou = 0.0 if union == 0 else intersect / union
        else:
            values = image_ious[class_id]
            class_iou = sum(values) / len(values) if values else None

        if class_iou is not None:
            global_ious.append(class_iou)

        image_values = image_ious[class_id]
        mean_image_iou = (
            sum(image_values) / len(image_values) if image_values else None
        )
        class_metrics.append(
            ClassSummary(
                class_id=class_id,
                class_name=class_name_for(class_id, class_names),
                iou=class_iou,
                mean_image_iou=mean_image_iou,
                intersect=intersect,
                pred_area=pred_area,
                label_area=label_area,
                union=union,
            )
        )

    summary = ModelSummary(
        model_name=model_name,
        csv_file=path.name,
        csv_path=str(path.resolve()),
        num_images=len(rows),
        miou=(sum(global_ious) / len(global_ious) if global_ious else None),
        mean_image_miou=(sum(image_mious) / len(image_mious) if image_mious else None),
        class_metrics=class_metrics,
    )
    summary.warnings = [issue.message for issue in issues if issue.level == "WARNING"]
    return summary, issues


def find_analysis_csvs(ana_dir: Path, recursive: bool) -> list[Path]:
    iterator = ana_dir.rglob("per_image_miou.csv") if recursive else ana_dir.glob("*/per_image_miou.csv")
    return sorted((p for p in iterator if p.is_file()), key=lambda p: str(p).lower())


def append_rows(ws, headers, rows):
    ws.append(headers)
    for row in rows:
        ws.append(row)


def add_table(ws, table_name: str) -> None:
    if ws.max_row < 2 or ws.max_column < 1:
        return
    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def style_sheet(ws, freeze: str = "A2", metric_start_col: Optional[int] = None) -> None:
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_gray = Side(style="thin", color="D9E2F3")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=thin_gray)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center")

    if metric_start_col is not None and ws.max_row >= 2:
        for row in ws.iter_rows(min_row=2, min_col=metric_start_col, max_col=ws.max_column):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = "0.0000"

        start = get_column_letter(metric_start_col)
        end = get_column_letter(ws.max_column)
        ws.conditional_formatting.add(
            f"{start}2:{end}{ws.max_row}",
            ColorScaleRule(
                start_type="min", start_color="F8696B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="63BE7B",
            ),
        )

    for column_cells in ws.columns:
        max_len = 0
        column = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[column].width = min(max(max_len + 2, 10), 60)


def collect_canonical_classes(records: Iterable[ModelSummary]) -> list[tuple[int, str]]:
    names: dict[int, str] = {}
    for record in records:
        for metric in record.class_metrics:
            names.setdefault(metric.class_id, metric.class_name)
    return sorted(names.items())


def safe_number(value):
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return None
    return value


def export_workbook(records: list[ModelSummary], issues: list[ParseIssue], output_path: Path) -> None:
    global Workbook
    global ColorScaleRule
    global Alignment
    global Border
    global Font
    global PatternFill
    global Side
    global get_column_letter
    global Table
    global TableStyleInfo

    try:
        from openpyxl import Workbook
        from openpyxl.formatting.rule import ColorScaleRule
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError as exc:  # pragma: no cover
        raise SystemExit(
            "Missing dependency: openpyxl. Install it with: pip install openpyxl"
        ) from exc

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"

    canonical_classes = collect_canonical_classes(records)

    summary_headers = [
        "Model",
        "CSV File",
        "Images",
        "mIoU",
        "Mean Image mIoU",
        "Warnings",
        "CSV Path",
    ]
    summary_rows = [
        [
            r.model_name,
            r.csv_file,
            r.num_images,
            safe_number(r.miou),
            safe_number(r.mean_image_miou),
            "; ".join(r.warnings),
            r.csv_path,
        ]
        for r in records
    ]
    append_rows(summary, summary_headers, summary_rows)
    style_sheet(summary, freeze="A2", metric_start_col=4)
    add_table(summary, "AnalysisSummaryTable")

    iou_ws = wb.create_sheet("Class IoU")
    iou_headers = ["Model", "Images", "mIoU"] + [
        name for _class_id, name in canonical_classes
    ]
    iou_rows = []
    for record in records:
        by_id = {metric.class_id: metric for metric in record.class_metrics}
        iou_rows.append([
            record.model_name,
            record.num_images,
            safe_number(record.miou),
            *[
                safe_number(by_id[class_id].iou) if class_id in by_id else None
                for class_id, _name in canonical_classes
            ],
        ])
    append_rows(iou_ws, iou_headers, iou_rows)
    style_sheet(iou_ws, freeze="D2", metric_start_col=3)
    add_table(iou_ws, "ClassIoUTable")

    image_iou_ws = wb.create_sheet("Mean Image Class IoU")
    image_iou_headers = ["Model", "Images", "Mean Image mIoU"] + [
        name for _class_id, name in canonical_classes
    ]
    image_iou_rows = []
    for record in records:
        by_id = {metric.class_id: metric for metric in record.class_metrics}
        image_iou_rows.append([
            record.model_name,
            record.num_images,
            safe_number(record.mean_image_miou),
            *[
                safe_number(by_id[class_id].mean_image_iou) if class_id in by_id else None
                for class_id, _name in canonical_classes
            ],
        ])
    append_rows(image_iou_ws, image_iou_headers, image_iou_rows)
    style_sheet(image_iou_ws, freeze="D2", metric_start_col=3)
    add_table(image_iou_ws, "MeanImageClassIoUTable")

    details = wb.create_sheet("Per-Class Details")
    detail_headers = [
        "Model",
        "Class ID",
        "Class Name",
        "IoU",
        "Mean Image IoU",
        "Intersect",
        "Pred Area",
        "Label Area",
        "Union",
        "Images",
        "CSV Path",
    ]
    detail_rows = []
    for record in records:
        for metric in sorted(record.class_metrics, key=lambda item: item.class_id):
            detail_rows.append([
                record.model_name,
                metric.class_id,
                metric.class_name,
                safe_number(metric.iou),
                safe_number(metric.mean_image_iou),
                metric.intersect,
                metric.pred_area,
                metric.label_area,
                metric.union,
                record.num_images,
                record.csv_path,
            ])
    append_rows(details, detail_headers, detail_rows)
    style_sheet(details, freeze="A2", metric_start_col=4)
    add_table(details, "PerClassDetailsTable")

    issue_sheet = wb.create_sheet("Parse Issues")
    issue_headers = ["Source", "Level", "Message"]
    issue_rows = [[i.source, i.level, i.message] for i in issues]
    if not issue_rows:
        issue_rows = [["", "INFO", "No parsing issues detected"]]
    append_rows(issue_sheet, issue_headers, issue_rows)
    style_sheet(issue_sheet, freeze="A2")
    add_table(issue_sheet, "ParseIssuesTable")

    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["G"].width = 70
    details.column_dimensions["C"].width = 28
    details.column_dimensions["K"].width = 70
    issue_sheet.column_dimensions["C"].width = 80

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Summarize ana/<model>/per_image_miou.csv into one IoU workbook."
    )
    parser.add_argument(
        "ana_dir",
        type=Path,
        nargs="?",
        default=Path("ana"),
        help="Analysis root directory. Default: ana",
    )
    parser.add_argument(
        "-o", "--output",
        type=Path,
        default=None,
        help="Output .xlsx path; default: <ana_dir>/iou_summary.xlsx",
    )
    parser.add_argument(
        "-c", "--class-file",
        type=Path,
        default=None,
        help="Optional class-name text file, one class per line",
    )
    parser.add_argument(
        "-r", "--recursive",
        action="store_true",
        help="Search per_image_miou.csv recursively in ana_dir",
    )
    return parser


def main() -> int:
    args = build_parser().parse_args()

    if not args.ana_dir.is_dir():
        print(f"ERROR: analysis directory does not exist: {args.ana_dir}", file=sys.stderr)
        return 2

    output = args.output or (args.ana_dir / "iou_summary.xlsx")
    if output.suffix.lower() != ".xlsx":
        output = output.with_suffix(".xlsx")

    csv_paths = find_analysis_csvs(args.ana_dir, args.recursive)
    if not csv_paths:
        print(
            f"ERROR: no per_image_miou.csv files found in {args.ana_dir}",
            file=sys.stderr,
        )
        return 3

    class_names, issues = load_class_names(args.class_file)
    records: list[ModelSummary] = []

    for csv_path in csv_paths:
        record, parse_issues = parse_per_image_csv(csv_path, class_names)
        issues.extend(parse_issues)
        if record is not None:
            records.append(record)

    if not records:
        print("ERROR: no valid per_image_miou.csv files were parsed.", file=sys.stderr)
        for issue in issues:
            print(f"[{issue.level}] {issue.source}: {issue.message}", file=sys.stderr)
        return 4

    export_workbook(records, issues, output)

    print(f"Scanned models     : {len(csv_paths)}")
    print(f"Exported summaries : {len(records)}")
    print(f"Class names        : {len(class_names) if class_names else 'class indexes'}")
    print(f"Output             : {output.resolve()}")
    if issues:
        print(f"Issues/warnings    : {len(issues)} (see 'Parse Issues' sheet)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
