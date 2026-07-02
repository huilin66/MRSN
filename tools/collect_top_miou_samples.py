#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Collect top-N samples by an Image mIoU column.

The script reads ana/iou_summary.xlsx, selects images with the largest mIoU in
one model column, then creates per-sample folders containing:
  - generated RGB image from image1_path
  - generated GT pseudo-color image from label_path
  - CAM overlays found in each model folder
  - pred_color images from every model's per_image_miou.csv

Example:
    python tools/collect_top_miou_samples.py ana \
        --rank-column cxup_4b_BW_PMRG_v2_lossV2 \
        --topk 20 \
        --output-dir ana/top20_cxup_4b_BW_PMRG_v2_lossV2
"""

from __future__ import annotations

import argparse
import csv
import math
import re
import shutil
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional


DEFAULT_COLLECT_MODELS = [
    "cxup_1b_BW",
    "unet_BW",
    "deeplabv3p_BW",
    "ocrnet_BW",
    "segformer_BW",
    "highdan_BW",
    "cxup_4b2h_BW",
    "cxup_4b_BW_PMRG_v2_lossV2",
]

BRIGHT_COLORS = [
    (0, 0, 0),
    (180, 180, 180),
    (60, 180, 75),
    (255, 225, 25),
    (0, 130, 200),
    (245, 130, 48),
    (145, 30, 180),
    (70, 240, 240),
    (240, 50, 230),
    (210, 245, 60),
    (250, 190, 190),
    (0, 128, 128),
    (230, 190, 255),
    (170, 110, 40),
    (255, 250, 200),
    (128, 0, 0),
    (170, 255, 195),
    (128, 128, 0),
    (255, 215, 180),
    (0, 0, 128),
]


@dataclass
class SelectedSample:
    rank: int
    image_key: str
    index: Optional[int]
    miou: float
    baseline_miou: Optional[float]
    miou_gain: Optional[float]
    image1_path: str
    image2_path: str
    label_path: str


@dataclass
class ModelAssets:
    model_name: str
    pred_gray_by_key: dict[str, str] = field(default_factory=dict)
    pred_color_by_key: dict[str, str] = field(default_factory=dict)
    cam_by_key: dict[str, str] = field(default_factory=dict)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Collect RGB/GT/CAM/pred_color assets for top-N mIoU samples."
    )
    parser.add_argument(
        "ana_dir",
        type=Path,
        nargs="?",
        default=Path("ana"),
        help="Analysis root directory. Default: ana",
    )
    parser.add_argument(
        "--summary",
        type=Path,
        default=None,
        help="Summary xlsx path. Default: <ana_dir>/iou_summary.xlsx",
    )
    parser.add_argument(
        "--rank-column",
        required=True,
        help="Model column in the 'Image mIoU' sheet used for ranking.",
    )
    parser.add_argument(
        "--topk",
        type=int,
        default=20,
        help="Number of samples to collect. Default: 20",
    )
    parser.add_argument(
        "--min-miou",
        type=float,
        default=0.5,
        help="Only collect samples whose rank-column mIoU is greater than this value. Default: 0.5",
    )
    parser.add_argument(
        "--better-than-column",
        default="cxup_1b_BW",
        help=(
            "Only collect samples whose rank-column mIoU is greater than this comparison column. "
            "Default: cxup_1b_BW"
        ),
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=None,
        help="Output directory. Default: <ana_dir>/top<k>_<rank-column>",
    )
    parser.add_argument(
        "--rgb-bands",
        nargs=3,
        type=int,
        default=[0, 1, 2],
        help="Zero-based bands used to generate RGB from image1_path. Default: 0 1 2",
    )
    parser.add_argument(
        "--sheet",
        default="Image mIoU",
        help="Workbook sheet name for ranking. Default: Image mIoU",
    )
    parser.add_argument(
        "--no-cam",
        action="store_true",
        help="Do not copy CAM overlays.",
    )
    parser.add_argument(
        "--no-pred",
        action="store_true",
        help="Do not copy model pred_color images.",
    )
    parser.add_argument(
        "--models",
        nargs="+",
        default=DEFAULT_COLLECT_MODELS,
        help=(
            "Model folders to collect. Default: "
            + ", ".join(DEFAULT_COLLECT_MODELS)
        ),
    )
    parser.add_argument(
        "--all-models",
        action="store_true",
        help="Collect every model folder under ana_dir instead of the default subset.",
    )
    parser.add_argument(
        "--path-map",
        nargs=2,
        action="append",
        metavar=("FROM", "TO"),
        default=[],
        help=(
            "Map source data paths before reading RGB/GT, for example: "
            "--path-map /data/huilin/data/C2Seg_BW/train E:\\data\\C2Seg_BW\\train"
        ),
    )
    parser.add_argument(
        "--class-file",
        type=Path,
        default=None,
        help="Optional class-name text file used to generate legend.png.",
    )
    parser.add_argument(
        "--num-classes",
        type=int,
        default=14,
        help="Number of classes for legend when class-file is not provided. Default: 14",
    )
    return parser


def safe_float(value) -> Optional[float]:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        result = float(text)
    except ValueError:
        return None
    if math.isnan(result) or math.isinf(result):
        return None
    return result


def optional_int(value) -> Optional[int]:
    number = safe_float(value)
    return int(number) if number is not None else None


def sanitize_name(name: str) -> str:
    invalid = '<>:"/\\|?*'
    cleaned = "".join("_" if ch in invalid else ch for ch in str(name))
    cleaned = cleaned.strip().strip(".")
    return cleaned or "item"


def read_image_miou_sheet(summary_path: Path,
                          sheet_name: str,
                          rank_column: str,
                          topk: int,
                          min_miou: Optional[float],
                          better_than_column: Optional[str]):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise SystemExit(
            "Missing dependency: openpyxl. Install it with: pip install openpyxl"
        ) from exc

    wb = load_workbook(summary_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            "Sheet '{}' not found in {}. Available sheets: {}".format(
                sheet_name, summary_path, ", ".join(wb.sheetnames)
            )
        )

    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    if rank_column not in headers:
        model_columns = headers[5:]
        raise ValueError(
            "Rank column '{}' not found. Available model columns: {}".format(
                rank_column, ", ".join(model_columns)
            )
        )
    if better_than_column and better_than_column not in headers:
        model_columns = headers[5:]
        raise ValueError(
            "Comparison column '{}' not found. Available model columns: {}".format(
                better_than_column, ", ".join(model_columns)
            )
        )

    col = {name: idx for idx, name in enumerate(headers)}
    required = ["Image Key", "Index", "Image1 Path", "Image2 Path", "Label Path"]
    missing = [name for name in required if name not in col]
    if missing:
        raise ValueError("Missing required columns in '{}': {}".format(sheet_name, missing))

    samples = []
    for row in rows:
        miou = safe_float(row[col[rank_column]])
        if miou is None:
            continue
        if min_miou is not None and miou <= min_miou:
            continue
        if better_than_column:
            baseline_miou = safe_float(row[col[better_than_column]])
            if baseline_miou is None or miou <= baseline_miou:
                continue
        else:
            baseline_miou = None
        image_key = str(row[col["Image Key"]])
        samples.append(
            SelectedSample(
                rank=0,
                image_key=image_key,
                index=optional_int(row[col["Index"]]),
                miou=miou,
                baseline_miou=baseline_miou,
                miou_gain=(miou - baseline_miou) if baseline_miou is not None else None,
                image1_path=str(row[col["Image1 Path"]] or ""),
                image2_path=str(row[col["Image2 Path"]] or ""),
                label_path=str(row[col["Label Path"]] or ""),
            )
        )

    samples = sorted(samples, key=lambda item: item.miou, reverse=True)[:topk]
    for rank, sample in enumerate(samples, start=1):
        sample.rank = rank
    return samples


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def image_key_from_row(row: dict[str, str], row_index: int) -> str:
    index = optional_int(row.get("index"))
    if index is not None:
        return str(index)
    return row.get("label_path") or row.get("image1_path") or f"row_{row_index}"


def index_from_key(image_key: str) -> Optional[int]:
    try:
        return int(float(image_key))
    except ValueError:
        return None


def load_model_assets(ana_dir: Path, include_models: Optional[list[str]]) -> list[ModelAssets]:
    include = set(include_models) if include_models else None
    assets = []
    for model_dir in sorted([p for p in ana_dir.iterdir() if p.is_dir()], key=lambda p: p.name.lower()):
        if include is not None and model_dir.name not in include:
            continue

        model = ModelAssets(model_name=model_dir.name)

        per_image = model_dir / "per_image_miou.csv"
        if per_image.is_file():
            for row_index, row in enumerate(read_csv(per_image)):
                key = image_key_from_row(row, row_index)
                pred_gray = row.get("pred_gray_path") or ""
                pred_color = row.get("pred_color_path") or ""
                if pred_gray:
                    model.pred_gray_by_key[key] = pred_gray
                if pred_color:
                    model.pred_color_by_key[key] = pred_color

        cam_meta = model_dir / "cam_meta.csv"
        if cam_meta.is_file():
            for row_index, row in enumerate(read_csv(cam_meta)):
                key = image_key_from_row(row, row_index)
                cam_path = row.get("cam_overlay_path") or ""
                if cam_path:
                    model.cam_by_key[key] = cam_path

        assets.append(model)
    return assets


def resolve_existing_path(path_text: str, ana_dir: Path, model_dir: Optional[Path] = None) -> Optional[Path]:
    if not path_text:
        return None

    raw = Path(path_text)
    candidates = []
    if raw.is_absolute():
        candidates.append(raw)
    else:
        candidates.append(Path.cwd() / raw)
        candidates.append(ana_dir / raw)
        if model_dir is not None:
            candidates.append(model_dir / raw)
    for candidate in candidates:
        if candidate.is_file():
            return candidate
    return None


def apply_path_maps(path_text: str, path_maps: list[list[str]]) -> str:
    normalized = str(path_text).replace("\\", "/")
    for src, dst in path_maps:
        src_norm = str(src).replace("\\", "/").rstrip("/")
        if normalized == src_norm or normalized.startswith(src_norm + "/"):
            suffix = normalized[len(src_norm):].lstrip("/")
            return str(Path(dst) / Path(*suffix.split("/"))) if suffix else str(Path(dst))
    return str(path_text)


def resolve_data_path(path_text: str, path_maps: list[list[str]]) -> Path:
    mapped = apply_path_maps(path_text, path_maps)
    path = Path(mapped)
    if path.is_file():
        return path

    if not path.is_absolute():
        cwd_path = Path.cwd() / path
        if cwd_path.is_file():
            return cwd_path

    raise FileNotFoundError(str(path))


def prediction_file_name(sample: SelectedSample) -> str:
    source = sample.label_path or sample.image1_path or sample.image_key
    stem = Path(source).stem
    sample_index = sample.index if sample.index is not None else index_from_key(sample.image_key)
    if sample_index is None:
        return "{}.png".format(sanitize_name(stem))
    return "{:06d}_{}.png".format(sample_index, stem)


def cam_file_name(sample: SelectedSample) -> str:
    source = sample.image1_path or sample.label_path or sample.image_key
    stem = Path(source).stem
    sample_index = sample.index if sample.index is not None else index_from_key(sample.image_key)
    if sample_index is None:
        return "{}.png".format(sanitize_name(stem))
    return "{:06d}_{}.png".format(sample_index, stem)


def fallback_pred_path(ana_dir: Path, model_name: str, sample: SelectedSample) -> Optional[Path]:
    path = ana_dir / model_name / "color" / prediction_file_name(sample)
    return path if path.is_file() else None


def fallback_pred_gray_path(ana_dir: Path, model_name: str, sample: SelectedSample) -> Optional[Path]:
    path = ana_dir / model_name / "gray" / prediction_file_name(sample)
    return path if path.is_file() else None


def fallback_cam_path(ana_dir: Path, model_name: str, sample: SelectedSample) -> Optional[Path]:
    path = ana_dir / model_name / "cam_overlay" / cam_file_name(sample)
    return path if path.is_file() else None


def bright_color_map(num_classes=256):
    palette = []
    for class_id in range(num_classes):
        if class_id < len(BRIGHT_COLORS):
            palette.extend(BRIGHT_COLORS[class_id])
        else:
            hue = (class_id * 47) % 360
            saturation = 0.82
            value = 1.0
            palette.extend(hsv_to_rgb_uint8(hue, saturation, value))
    return palette


def read_text_auto(path: Path) -> str:
    for encoding in ("utf-8", "utf-8-sig", "gb18030", "latin-1"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    return path.read_text(encoding="utf-8", errors="replace")


def load_class_names(path: Optional[Path], num_classes: int) -> list[str]:
    names = ["class_{}".format(i) for i in range(num_classes)]
    if path is None:
        return names
    if not path.is_file():
        raise FileNotFoundError("Class file does not exist: {}".format(path))

    loaded = []
    for raw_line in read_text_auto(path).splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        match = re.match(r"^\s*\d+\s*(?:[:;,|\t]|\s)\s*(.+?)\s*$", line)
        if match:
            line = match.group(1).strip()
        loaded.append(line)

    if len(loaded) > len(names):
        names.extend("class_{}".format(i) for i in range(len(names), len(loaded)))
    for class_id, class_name in enumerate(loaded):
        names[class_id] = class_name
    return names


def hsv_to_rgb_uint8(hue, saturation, value):
    c = value * saturation
    x = c * (1 - abs((hue / 60) % 2 - 1))
    m = value - c
    if hue < 60:
        r, g, b = c, x, 0
    elif hue < 120:
        r, g, b = x, c, 0
    elif hue < 180:
        r, g, b = 0, c, x
    elif hue < 240:
        r, g, b = 0, x, c
    elif hue < 300:
        r, g, b = x, 0, c
    else:
        r, g, b = c, 0, x
    return [int((r + m) * 255), int((g + m) * 255), int((b + m) * 255)]


def read_array(path: str):
    try:
        from skimage import io as skio
        return skio.imread(path)
    except Exception:
        from PIL import Image
        return np.asarray(Image.open(path))


def image_to_rgb(image, bands):
    image = np.asarray(image)
    if image.ndim == 2:
        image = np.stack([image, image, image], axis=-1)
    elif image.ndim == 3 and image.shape[0] <= 16 and image.shape[-1] > 16:
        image = np.transpose(image, (1, 2, 0))

    if image.ndim != 3:
        raise ValueError("Unsupported image shape for RGB extraction: {}".format(image.shape))

    max_band = image.shape[2] - 1
    selected = [min(max(band, 0), max_band) for band in bands]
    rgb = image[:, :, selected].astype("float32")

    channels = []
    for channel_idx in range(3):
        channel = rgb[:, :, channel_idx]
        finite = np.isfinite(channel)
        if not finite.any():
            channels.append(np.zeros_like(channel, dtype="uint8"))
            continue
        lo, hi = np.percentile(channel[finite], [2, 98])
        if hi <= lo:
            lo, hi = float(channel[finite].min()), float(channel[finite].max())
        if hi <= lo:
            scaled = np.zeros_like(channel, dtype="uint8")
        else:
            scaled = np.clip((channel - lo) / (hi - lo), 0, 1)
            scaled = (scaled * 255).astype("uint8")
        channels.append(scaled)
    return np.stack(channels, axis=-1)


def save_rgb_image(image_path: str, output_path: Path, bands):
    from PIL import Image

    image = read_array(image_path)
    rgb = image_to_rgb(image, bands)
    Image.fromarray(rgb).save(output_path)


def save_gt_color(label_path: str, output_path: Path):
    from PIL import Image

    label = read_array(label_path)
    label = np.asarray(label).squeeze().astype("uint8")
    label_img = Image.fromarray(label, mode="P")
    label_img.putpalette(bright_color_map(256))
    label_img.convert("RGB").save(output_path)


def save_mask_color(mask_path: Path, output_path: Path):
    from PIL import Image

    mask = read_array(str(mask_path))
    mask = np.asarray(mask).squeeze().astype("uint8")
    mask_img = Image.fromarray(mask, mode="P")
    mask_img.putpalette(bright_color_map(256))
    output_path.parent.mkdir(parents=True, exist_ok=True)
    mask_img.convert("RGB").save(output_path)


def copy_if_exists(src: Optional[Path], dst: Path) -> bool:
    if src is None or not src.is_file():
        return False
    dst.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, dst)
    return True


def save_legend(output_path: Path, class_names: list[str]):
    from PIL import Image, ImageDraw, ImageFont

    swatch = 28
    row_h = 36
    pad = 14
    text_x = pad + swatch + 12
    width = 360
    height = pad * 2 + row_h * len(class_names)
    colors = BRIGHT_COLORS

    image = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(image)
    try:
        font = ImageFont.truetype("arial.ttf", 16)
    except OSError:
        font = ImageFont.load_default()

    for class_id, class_name in enumerate(class_names):
        y = pad + class_id * row_h
        color = colors[class_id] if class_id < len(colors) else tuple(hsv_to_rgb_uint8((class_id * 47) % 360, 0.82, 1.0))
        draw.rectangle([pad, y + 4, pad + swatch, y + 4 + swatch], fill=color, outline=(80, 80, 80))
        draw.text((text_x, y + 8), "{}: {}".format(class_id, class_name), fill=(20, 20, 20), font=font)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    image.save(output_path)


def write_manifest(path: Path, rows: list[dict[str, object]]):
    if not rows:
        return
    fieldnames = sorted({key for row in rows for key in row.keys()})
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def collect_sample(sample: SelectedSample, ana_dir: Path, output_dir: Path, model_assets: list[ModelAssets], args):
    sample_name = "rank_{:02d}_idx_{}_miou_{:.4f}".format(
        sample.rank,
        sanitize_name(sample.image_key),
        sample.miou,
    )
    sample_dir = output_dir / sample_name
    sample_dir.mkdir(parents=True, exist_ok=True)

    manifest_rows = []

    rgb_path = sample_dir / "rgb.png"
    gt_color_path = sample_dir / "gt_color.png"
    try:
        resolved_rgb = resolve_data_path(sample.image1_path, args.path_map)
        save_rgb_image(str(resolved_rgb), rgb_path, args.rgb_bands)
        manifest_rows.append({"type": "rgb", "path": str(rgb_path), "source": str(resolved_rgb)})
    except Exception as exc:
        manifest_rows.append({"type": "rgb", "path": "", "source": sample.image1_path, "error": str(exc)})

    try:
        resolved_label = resolve_data_path(sample.label_path, args.path_map)
        save_gt_color(str(resolved_label), gt_color_path)
        manifest_rows.append({"type": "gt_color", "path": str(gt_color_path), "source": str(resolved_label)})
    except Exception as exc:
        manifest_rows.append({"type": "gt_color", "path": "", "source": sample.label_path, "error": str(exc)})

    for model in model_assets:
        model_dir = ana_dir / model.model_name
        key = sample.image_key

        if not args.no_pred:
            pred_dst = sample_dir / "pred_color" / "{}.png".format(sanitize_name(model.model_name))
            pred_gray_src = resolve_existing_path(model.pred_gray_by_key.get(key, ""), ana_dir, model_dir)
            if pred_gray_src is None:
                pred_gray_src = fallback_pred_gray_path(ana_dir, model.model_name, sample)

            copied = False
            pred_src = pred_gray_src
            error = ""
            if pred_gray_src is not None:
                try:
                    save_mask_color(pred_gray_src, pred_dst)
                    copied = True
                except Exception as exc:
                    error = str(exc)

            if not copied:
                pred_src = resolve_existing_path(model.pred_color_by_key.get(key, ""), ana_dir, model_dir)
                if pred_src is None:
                    pred_src = fallback_pred_path(ana_dir, model.model_name, sample)
                copied = copy_if_exists(pred_src, pred_dst)
            manifest_rows.append({
                "type": "pred_color",
                "model": model.model_name,
                "path": str(pred_dst) if copied else "",
                "source": str(pred_src) if pred_src else "",
                "found": copied,
                "error": error,
            })

        if not args.no_cam:
            cam_src = resolve_existing_path(model.cam_by_key.get(key, ""), ana_dir, model_dir)
            if cam_src is None:
                cam_src = fallback_cam_path(ana_dir, model.model_name, sample)
            cam_dst = sample_dir / "cam" / "{}.png".format(sanitize_name(model.model_name))
            copied = copy_if_exists(cam_src, cam_dst)
            manifest_rows.append({
                "type": "cam",
                "model": model.model_name,
                "path": str(cam_dst) if copied else "",
                "source": str(cam_src) if cam_src else "",
                "found": copied,
            })

    write_manifest(sample_dir / "manifest.csv", manifest_rows)
    return {
        "rank": sample.rank,
        "image_key": sample.image_key,
        "index": sample.index,
        "miou": "{:.8f}".format(sample.miou),
        "baseline_miou": "{:.8f}".format(sample.baseline_miou) if sample.baseline_miou is not None else "",
        "miou_gain": "{:.8f}".format(sample.miou_gain) if sample.miou_gain is not None else "",
        "sample_dir": str(sample_dir),
        "image1_path": sample.image1_path,
        "image2_path": sample.image2_path,
        "label_path": sample.label_path,
    }


def main() -> int:
    args = build_parser().parse_args()

    global np
    import numpy as np

    ana_dir = args.ana_dir
    if not ana_dir.is_dir():
        print("ERROR: ana_dir does not exist: {}".format(ana_dir), file=sys.stderr)
        return 2

    summary_path = args.summary or (ana_dir / "iou_summary.xlsx")
    if not summary_path.is_file():
        print("ERROR: summary xlsx does not exist: {}".format(summary_path), file=sys.stderr)
        return 3

    output_dir = args.output_dir or (
        ana_dir / "top{}_{}".format(args.topk, sanitize_name(args.rank_column))
    )
    output_dir.mkdir(parents=True, exist_ok=True)
    try:
        class_names = load_class_names(args.class_file, args.num_classes)
        save_legend(output_dir / "legend.png", class_names)
    except Exception as exc:
        print("WARNING: failed to generate legend.png: {}".format(exc), file=sys.stderr)

    samples = read_image_miou_sheet(
        summary_path,
        args.sheet,
        args.rank_column,
        args.topk,
        args.min_miou,
        args.better_than_column,
    )
    if not samples:
        print("ERROR: no valid samples found for rank column {}".format(args.rank_column), file=sys.stderr)
        return 4

    include_models = None if args.all_models else args.models
    model_assets = load_model_assets(ana_dir, include_models)
    if not model_assets:
        print("ERROR: no matching model folders found in {}".format(ana_dir), file=sys.stderr)
        return 5
    summary_rows = []
    for sample in samples:
        summary_rows.append(collect_sample(sample, ana_dir, output_dir, model_assets, args))
        print("Collected rank {}/{}: image_key={}, mIoU={:.4f}".format(
            sample.rank, len(samples), sample.image_key, sample.miou
        ))

    write_manifest(output_dir / "top_samples.csv", summary_rows)
    print("Output: {}".format(output_dir.resolve()))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
