#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Batch parse validation .log files and export one comparison workbook.

Example:
    python logs_to_xlsx.py ./log/val \
        --class-file ./class.txt \
        --output ./val_results.xlsx \
        --recursive

Dependencies:
    pip install openpyxl
"""

from __future__ import annotations

import argparse
import math
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable, Optional

try:
    from openpyxl import Workbook
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "Missing dependency: openpyxl. Install it with: pip install openpyxl"
    ) from exc


FLOAT = r"[-+]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][-+]?\d+)?"

OVERALL_RE = re.compile(
    rf"^(?P<timestamp>\d{{4}}-\d{{2}}-\d{{2}}\s+\d{{2}}:\d{{2}}:\d{{2}}).*?"
    rf"\[EVAL\]\s*#Images:\s*(?P<images>\d+)\s+"
    rf"F1:\s*(?P<f1>{FLOAT})\s*,?\s*"
    rf"mIoU:\s*(?P<miou>{FLOAT})\s+"
    rf"Acc:\s*(?P<acc>{FLOAT})\s+"
    rf"Kappa:\s*(?P<kappa>{FLOAT})",
    re.MULTILINE,
)

COMPLEXITY_RE = re.compile(
    rf"\[EVAL\]\s*Params:\s*(?P<params>{FLOAT})M\s+"
    rf"Trainable:\s*(?P<trainable>{FLOAT})M\s+"
    rf"FLOPs:\s*(?P<flops>{FLOAT})G\s+"
    rf"FPS:\s*(?P<fps>{FLOAT})"
)

GPU_RE = re.compile(
    r"\[GPU Memory\]\s*Used:\s*(?P<used>\d+)\s*MiB\s*/\s*(?P<total>\d+)\s*MiB"
    r"\s*\|\s*Paddle allocated:\s*(?P<allocated>\d+)\s*MiB"
    r"\s*\|\s*peak:\s*(?P<allocated_peak>\d+)\s*MiB"
    r"\s*\|\s*Paddle reserved:\s*(?P<reserved>\d+)\s*MiB"
    r"\s*\|\s*peak:\s*(?P<reserved_peak>\d+)\s*MiB"
)

CLASS_ROW_RE = re.compile(
    rf"^\s*(?P<id>\d+)\s*\|\s*(?P<name>[^|]+?)\s*\|\s*"
    rf"(?P<iou>{FLOAT})\s*\|\s*(?P<f1>{FLOAT})\s*\|\s*"
    rf"(?P<acc>{FLOAT})\s*\|\s*(?P<intersect>\d+)\s*\|\s*"
    rf"(?P<pred>\d+)\s*\|\s*(?P<label>\d+)\s*$",
    re.MULTILINE,
)

ARRAY_PATTERNS = {
    "iou": re.compile(r"\[EVAL\]\s*Class IoU:\s*\n\s*\[([^\]]+)\]", re.DOTALL),
    "acc": re.compile(r"\[EVAL\]\s*Class Acc:\s*\n\s*\[([^\]]+)\]", re.DOTALL),
    "f1": re.compile(r"\[EVAL\]\s*Class F1:\s*\n\s*\[([^\]]+)\]", re.DOTALL),
}


@dataclass
class ClassMetric:
    class_id: int
    class_name: str
    iou: Optional[float] = None
    f1: Optional[float] = None
    acc: Optional[float] = None
    intersect: Optional[int] = None
    pred: Optional[int] = None
    label: Optional[int] = None


@dataclass
class EvalRecord:
    experiment: str
    log_file: str
    log_path: str
    eval_index: int
    timestamp: str
    model: Optional[str] = None
    backbone: Optional[str] = None
    checkpoint: Optional[str] = None
    configured_num_classes: Optional[int] = None
    images: Optional[int] = None
    f1: Optional[float] = None
    miou: Optional[float] = None
    acc: Optional[float] = None
    kappa: Optional[float] = None
    params_m: Optional[float] = None
    trainable_m: Optional[float] = None
    flops_g: Optional[float] = None
    fps: Optional[float] = None
    gpu_used_mib: Optional[int] = None
    gpu_total_mib: Optional[int] = None
    paddle_allocated_mib: Optional[int] = None
    paddle_allocated_peak_mib: Optional[int] = None
    paddle_reserved_mib: Optional[int] = None
    paddle_reserved_peak_mib: Optional[int] = None
    class_metrics: list[ClassMetric] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


@dataclass
class ParseIssue:
    log_file: str
    level: str
    message: str


def read_text_auto(path: Path) -> str:
    """Read common log encodings without failing the whole batch."""
    for encoding in ("utf-8", "utf-8-sig", "gb18030", "latin-1"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    return path.read_text(encoding="utf-8", errors="replace")


def to_float(value: Optional[str]) -> Optional[float]:
    return float(value) if value is not None else None


def to_int(value: Optional[str]) -> Optional[int]:
    return int(value) if value is not None else None


def first_group(pattern: str, text: str, flags: int = 0) -> Optional[str]:
    match = re.search(pattern, text, flags)
    return match.group(1).strip() if match else None


def parse_model_metadata(text: str) -> dict[str, object]:
    model_block_match = re.search(
        r"(?ms)^model:\s*\n(?P<block>.*?)(?:^[-]{10,}\s*$|\Z)", text
    )
    block = model_block_match.group("block") if model_block_match else ""

    model = first_group(r"(?m)^\s*type:\s*(\S+)\s*$", block)
    backbone = first_group(r"(?m)^\s*backb(?:one)?:\s*(\S+)\s*$", block)
    num_classes_text = first_group(r"(?m)^\s*num_classes:\s*(\d+)\s*$", block)

    checkpoint = first_group(
        r"(?m)Loading pretrained model from\s+(.+?)\s*$", text
    )

    # Fallback if the model config block is absent.
    if model is None:
        model = first_group(
            r"(?m)variables loaded into\s+([A-Za-z0-9_.-]+)\.?\s*$", text
        )

    log_declared = first_group(r"(?m)^.*?Log file:\s*(.+?)\s*$", text)

    return {
        "model": model,
        "backbone": backbone,
        "num_classes": int(num_classes_text) if num_classes_text else None,
        "checkpoint": checkpoint,
        "log_declared": log_declared,
    }


def infer_experiment(path: Path, declared_log: Optional[str]) -> str:
    source = Path(declared_log).stem if declared_log else path.stem
    # Remove a conventional trailing validation timestamp.
    return re.sub(r"_\d{8}_\d{6}$", "", source)


def parse_numeric_array(block: str, metric: str) -> list[float]:
    match = ARRAY_PATTERNS[metric].search(block)
    if not match:
        return []
    return [float(item) for item in re.findall(FLOAT, match.group(1))]


def parse_class_metrics(block: str) -> list[ClassMetric]:
    table_rows = list(CLASS_ROW_RE.finditer(block))
    if table_rows:
        return [
            ClassMetric(
                class_id=int(m.group("id")),
                class_name=m.group("name").strip(),
                iou=float(m.group("iou")),
                f1=float(m.group("f1")),
                acc=float(m.group("acc")),
                intersect=int(m.group("intersect")),
                pred=int(m.group("pred")),
                label=int(m.group("label")),
            )
            for m in table_rows
        ]

    # Fallback for logs containing arrays but not the per-class table.
    ious = parse_numeric_array(block, "iou")
    f1s = parse_numeric_array(block, "f1")
    accs = parse_numeric_array(block, "acc")
    count = max(len(ious), len(f1s), len(accs), 0)

    return [
        ClassMetric(
            class_id=i,
            class_name=f"class_{i}",
            iou=ious[i] if i < len(ious) else None,
            f1=f1s[i] if i < len(f1s) else None,
            acc=accs[i] if i < len(accs) else None,
        )
        for i in range(count)
    ]


def parse_log(path: Path, eval_mode: str) -> tuple[list[EvalRecord], list[ParseIssue]]:
    text = read_text_auto(path)
    metadata = parse_model_metadata(text)
    experiment = infer_experiment(path, metadata.get("log_declared"))

    matches = list(OVERALL_RE.finditer(text))
    issues: list[ParseIssue] = []

    if not matches:
        issues.append(ParseIssue(path.name, "ERROR", "No [EVAL] overall metric line found"))
        return [], issues

    selected_indices = range(len(matches)) if eval_mode == "all" else [len(matches) - 1]
    records: list[EvalRecord] = []

    for output_index, source_index in enumerate(selected_indices, start=1):
        overall = matches[source_index]
        next_start = matches[source_index + 1].start() if source_index + 1 < len(matches) else len(text)
        block = text[overall.start():next_start]

        complexity = COMPLEXITY_RE.search(block)
        gpu = GPU_RE.search(block)
        class_metrics = parse_class_metrics(block)

        record = EvalRecord(
            experiment=experiment,
            log_file=path.name,
            log_path=str(path.resolve()),
            eval_index=(source_index + 1),
            timestamp=overall.group("timestamp"),
            model=metadata.get("model"),
            backbone=metadata.get("backbone"),
            checkpoint=metadata.get("checkpoint"),
            configured_num_classes=metadata.get("num_classes"),
            images=int(overall.group("images")),
            f1=float(overall.group("f1")),
            miou=float(overall.group("miou")),
            acc=float(overall.group("acc")),
            kappa=float(overall.group("kappa")),
            params_m=to_float(complexity.group("params")) if complexity else None,
            trainable_m=to_float(complexity.group("trainable")) if complexity else None,
            flops_g=to_float(complexity.group("flops")) if complexity else None,
            fps=to_float(complexity.group("fps")) if complexity else None,
            gpu_used_mib=to_int(gpu.group("used")) if gpu else None,
            gpu_total_mib=to_int(gpu.group("total")) if gpu else None,
            paddle_allocated_mib=to_int(gpu.group("allocated")) if gpu else None,
            paddle_allocated_peak_mib=to_int(gpu.group("allocated_peak")) if gpu else None,
            paddle_reserved_mib=to_int(gpu.group("reserved")) if gpu else None,
            paddle_reserved_peak_mib=to_int(gpu.group("reserved_peak")) if gpu else None,
            class_metrics=class_metrics,
        )

        if complexity is None:
            record.warnings.append("Complexity line not found")
        if gpu is None:
            record.warnings.append("GPU memory line not found")
        if not class_metrics:
            record.warnings.append("Per-class metrics not found")
        elif record.configured_num_classes is not None and len(class_metrics) != record.configured_num_classes:
            record.warnings.append(
                f"Parsed {len(class_metrics)} classes, but config declares "
                f"{record.configured_num_classes}"
            )

        for warning in record.warnings:
            issues.append(ParseIssue(path.name, "WARNING", warning))
        records.append(record)

    return records, issues


def load_class_names(path: Optional[Path]) -> tuple[list[str], list[ParseIssue]]:
    if path is None:
        return [], []
    if not path.is_file():
        return [], [ParseIssue(path.name, "ERROR", f"Class file does not exist: {path}")]

    lines = []
    for raw_line in read_text_auto(path).splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue

        # Also accept formats such as "0: Background", "0,Background", "0 Background".
        match = re.match(r"^\s*\d+\s*(?:[:;,|\t]|\s)\s*(.+?)\s*$", line)
        if match:
            line = match.group(1).strip()
        lines.append(line)

    return lines, []


def apply_class_names(
    records: list[EvalRecord], class_names: list[str]
) -> list[ParseIssue]:
    issues: list[ParseIssue] = []
    if not class_names:
        return issues

    max_class_count = max((len(r.class_metrics) for r in records), default=0)
    if len(class_names) != max_class_count:
        issues.append(
            ParseIssue(
                "class file",
                "WARNING",
                f"Class file contains {len(class_names)} names, while parsed logs contain "
                f"up to {max_class_count} classes. Matching IDs will be replaced; "
                "unmatched classes keep their original names.",
            )
        )

    for record in records:
        for metric in record.class_metrics:
            if 0 <= metric.class_id < len(class_names):
                metric.class_name = class_names[metric.class_id]

    return issues


def collect_canonical_classes(records: Iterable[EvalRecord]) -> list[tuple[int, str]]:
    names: dict[int, str] = {}
    for record in records:
        for metric in record.class_metrics:
            names.setdefault(metric.class_id, metric.class_name)
    return sorted(names.items())


def safe_number(value):
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return None
    return value


def add_table(ws, table_name: str) -> None:
    if ws.max_row < 2 or ws.max_column < 1:
        return
    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def style_sheet(ws, freeze: str = "A2", metric_start_col: Optional[int] = None) -> None:
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_gray = Side(style="thin", color="D9E2F3")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin_gray)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False)

    # Number formatting inferred from header names.
    for col_idx, cell in enumerate(ws[1], start=1):
        header = str(cell.value or "")
        if header in {"F1", "mIoU", "Acc", "Kappa", "IoU", "Class F1", "Class Acc"}:
            for data_cell in ws.iter_cols(
                min_col=col_idx, max_col=col_idx, min_row=2, max_row=ws.max_row
            ):
                for c in data_cell:
                    c.number_format = "0.0000"
        elif header in {"Params (M)", "Trainable (M)", "FLOPs (G)", "FPS"}:
            for data_cell in ws.iter_cols(
                min_col=col_idx, max_col=col_idx, min_row=2, max_row=ws.max_row
            ):
                for c in data_cell:
                    c.number_format = "0.00"

    # Sensible capped widths.
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter]:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 32)

    if metric_start_col is not None and ws.max_row >= 2 and ws.max_column >= metric_start_col:
        start = get_column_letter(metric_start_col)
        end = get_column_letter(ws.max_column)
        ws.conditional_formatting.add(
            f"{start}2:{end}{ws.max_row}",
            ColorScaleRule(
                start_type="min", start_color="F8696B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="63BE7B",
            ),
        )


def append_rows(ws, headers: list[str], rows: list[list[object]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([safe_number(v) for v in row])


def export_workbook(
    records: list[EvalRecord], issues: list[ParseIssue], output_path: Path
) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    canonical_classes = collect_canonical_classes(records)

    # 1. Overall summary.
    summary = wb.create_sheet("Summary")
    summary_headers = [
        "Experiment", "Log File", "Eval Index", "Timestamp", "Model", "Backbone",
        "Checkpoint", "Configured Classes", "Parsed Classes", "Images",
        "F1", "mIoU", "Acc", "Kappa", "Params (M)", "Trainable (M)",
        "FLOPs (G)", "FPS", "GPU Used (MiB)", "GPU Total (MiB)",
        "Paddle Allocated (MiB)", "Paddle Allocated Peak (MiB)",
        "Paddle Reserved (MiB)", "Paddle Reserved Peak (MiB)",
        "Warnings", "Log Path",
    ]
    summary_rows = []
    for r in records:
        summary_rows.append([
            r.experiment, r.log_file, r.eval_index, r.timestamp, r.model, r.backbone,
            r.checkpoint, r.configured_num_classes, len(r.class_metrics), r.images,
            r.f1, r.miou, r.acc, r.kappa, r.params_m, r.trainable_m,
            r.flops_g, r.fps, r.gpu_used_mib, r.gpu_total_mib,
            r.paddle_allocated_mib, r.paddle_allocated_peak_mib,
            r.paddle_reserved_mib, r.paddle_reserved_peak_mib,
            "; ".join(r.warnings), r.log_path,
        ])
    append_rows(summary, summary_headers, summary_rows)
    style_sheet(summary, freeze="A2", metric_start_col=11)
    add_table(summary, "SummaryTable")

    # 2-4. Wide matrices for IoU/F1/Acc.
    matrix_specs = [
        ("Class IoU", "iou", "ClassIoUTable"),
        ("Class F1", "f1", "ClassF1Table"),
        ("Class Acc", "acc", "ClassAccTable"),
    ]
    for sheet_name, attribute, table_name in matrix_specs:
        ws = wb.create_sheet(sheet_name)
        headers = ["Experiment", "Log File", "Eval Index", "Timestamp", "Model"] + [
            name for _, name in canonical_classes
        ]
        rows = []
        for r in records:
            by_id = {m.class_id: m for m in r.class_metrics}
            rows.append([
                r.experiment, r.log_file, r.eval_index, r.timestamp, r.model,
                *[
                    getattr(by_id[class_id], attribute) if class_id in by_id else None
                    for class_id, _ in canonical_classes
                ],
            ])
        append_rows(ws, headers, rows)
        style_sheet(ws, freeze="F2", metric_start_col=6)
        add_table(ws, table_name)

    # 5. Long-format per-class details.
    details = wb.create_sheet("Per-Class Details")
    detail_headers = [
        "Experiment", "Log File", "Eval Index", "Timestamp", "Model",
        "Class ID", "Class Name", "IoU", "F1", "Acc",
        "Intersect", "Pred", "Label",
    ]
    detail_rows = []
    for r in records:
        for m in sorted(r.class_metrics, key=lambda x: x.class_id):
            detail_rows.append([
                r.experiment, r.log_file, r.eval_index, r.timestamp, r.model,
                m.class_id, m.class_name, m.iou, m.f1, m.acc,
                m.intersect, m.pred, m.label,
            ])
    append_rows(details, detail_headers, detail_rows)
    style_sheet(details, freeze="A2", metric_start_col=8)
    add_table(details, "PerClassDetailsTable")

    # 6. Parse issues and warnings.
    issue_sheet = wb.create_sheet("Parse Issues")
    issue_headers = ["Log File", "Level", "Message"]
    issue_rows = [[i.log_file, i.level, i.message] for i in issues]
    if not issue_rows:
        issue_rows = [["", "INFO", "No parsing issues detected"]]
    append_rows(issue_sheet, issue_headers, issue_rows)
    style_sheet(issue_sheet, freeze="A2")
    add_table(issue_sheet, "ParseIssuesTable")

    # Make important summary columns easier to read.
    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["B"].width = 40
    summary.column_dimensions["G"].width = 55
    summary.column_dimensions["Y"].width = 36
    summary.column_dimensions["Z"].width = 60
    details.column_dimensions["G"].width = 38
    issue_sheet.column_dimensions["C"].width = 80

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def find_logs(log_dir: Path, pattern: str, recursive: bool) -> list[Path]:
    iterator = log_dir.rglob(pattern) if recursive else log_dir.glob(pattern)
    return sorted((p for p in iterator if p.is_file()), key=lambda p: str(p).lower())


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Parse validation log files and export one XLSX comparison workbook."
    )
    parser.add_argument("log_dir", type=Path, help="Folder containing .log files")
    parser.add_argument(
        "-o", "--output", type=Path, default=None,
        help="Output .xlsx path; default: <log_dir>/val_log_comparison.xlsx",
    )
    parser.add_argument(
        "-c", "--class-file", type=Path, default=None,
        help="Optional class-name text file, one class per line",
    )
    parser.add_argument(
        "-r", "--recursive", action="store_true",
        help="Search log files recursively in subfolders",
    )
    parser.add_argument(
        "--pattern", default="*.log",
        help="Log filename pattern, default: *.log",
    )
    parser.add_argument(
        "--eval-mode", choices=("all", "last"), default="all",
        help="Export all EVAL blocks or only the last block in each log; default: all",
    )
    return parser


def main() -> int:
    args = build_parser().parse_args()

    if not args.log_dir.is_dir():
        print(f"ERROR: log directory does not exist: {args.log_dir}", file=sys.stderr)
        return 2

    output = args.output or (args.log_dir / "val_log_comparison.xlsx")
    if output.suffix.lower() != ".xlsx":
        output = output.with_suffix(".xlsx")

    log_paths = find_logs(args.log_dir, args.pattern, args.recursive)
    if not log_paths:
        print(
            f"ERROR: no log files matched {args.pattern!r} in {args.log_dir}",
            file=sys.stderr,
        )
        return 3

    all_records: list[EvalRecord] = []
    all_issues: list[ParseIssue] = []

    for log_path in log_paths:
        records, issues = parse_log(log_path, args.eval_mode)
        all_records.extend(records)
        all_issues.extend(issues)

    class_names, class_issues = load_class_names(args.class_file)
    all_issues.extend(class_issues)
    all_issues.extend(apply_class_names(all_records, class_names))

    if not all_records:
        print("ERROR: no valid validation results were parsed.", file=sys.stderr)
        for issue in all_issues:
            print(f"[{issue.level}] {issue.log_file}: {issue.message}", file=sys.stderr)
        return 4

    export_workbook(all_records, all_issues, output)

    print(f"Scanned log files : {len(log_paths)}")
    print(f"Exported eval rows: {len(all_records)}")
    print(f"Class names       : {len(class_names) if class_names else 'from logs'}")
    print(f"Output            : {output.resolve()}")
    if all_issues:
        print(f"Issues/warnings   : {len(all_issues)} (see 'Parse Issues' sheet)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
